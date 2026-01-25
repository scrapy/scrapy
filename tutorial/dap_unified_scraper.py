"""
DAP RRI Unified Scraper - Combined Portal and Streaming Data Scraper

This script uses Playwright to scrape both:
1. Portal (Weblytic) data: PAGEVIEW, SESSION, NEW USERS, RETURNING USERS + Rankings
2. Streaming data: PRO 1-4 hits, ALL PRO, TOTAL HITS *DAP VER., TOTAL USER

Both datasets are exported to separate Excel files matching the exact template formats.

Usage:
    python dap_unified_scraper.py

    # With custom date range:
    python dap_unified_scraper.py --start-date 01/01/2026 --end-date 31/01/2026

    # Scrape only portal or streaming:
    python dap_unified_scraper.py --portal-only
    python dap_unified_scraper.py --streaming-only
"""

import asyncio
import argparse
from datetime import datetime
from playwright.async_api import async_playwright
import pandas as pd


# Configuration
EMAIL = "admin@rri.go.id"
PASSWORD = "!@#$ADMIN4daprri!@#$"
BASE_URL = "https://dap.rri.go.id"
PORTAL_URL = f"{BASE_URL}/portal/detail?page=6"  # Traffic Evolution page
STREAMING_URL = f"{BASE_URL}/streaming/detail?page=2"


async def login(page):
    """Login to DAP RRI."""
    print("[*] Logging in...")
    await page.goto(BASE_URL)
    await page.wait_for_load_state("networkidle")

    await page.fill('input[name="email"]', EMAIL)
    await page.fill('input[name="password"]', PASSWORD)
    await page.click("#kt_sign_in_submit")

    await page.wait_for_timeout(3000)
    print("[+] Login successful!")


async def get_satker_options(page):
    """Get all satker options from dropdown."""
    return await page.evaluate("""
        () => {
            const select = document.querySelector('#kt_dash_satker');
            if (select) {
                return Array.from(select.options).map(opt => ({
                    value: opt.value,
                    text: opt.text.trim()
                })).filter(opt => opt.value && opt.value !== '');
            }
            return [];
        }
    """)


async def set_date_range(page, start_date, end_date):
    """Set the date range filter."""
    date_range = f"{start_date} 00:00 - {end_date} 23:59"
    await page.evaluate(f'''
        () => {{
            const dateInput = document.querySelector('#kt_dash_daterangepicker');
            if (dateInput) {{
                dateInput.value = "{date_range}";
                dateInput.dispatchEvent(new Event('change'));
            }}
        }}
    ''')
    print(f"[*] Date range set: {date_range}")


async def scrape_portal_data(page, start_date, end_date):
    """Scrape Portal (Weblytic) data for all satkers."""
    print("\n" + "=" * 60)
    print("[*] PORTAL (WEBLYTIC) DATA SCRAPER")
    print("=" * 60)

    await page.goto(PORTAL_URL)
    await page.wait_for_load_state("networkidle")
    await page.wait_for_timeout(2000)

    satker_options = await get_satker_options(page)
    print(f"[*] Found {len(satker_options)} satkers to scrape")

    await set_date_range(page, start_date, end_date)

    results = []
    for idx, satker in enumerate(satker_options):
        satker_name = satker["text"]
        satker_value = satker["value"]

        print(f"[{idx + 1}/{len(satker_options)}] Scraping {satker_name}...")

        await page.select_option("#kt_dash_satker", value=satker_value)
        await page.click("#kt_apply_filter")
        await page.wait_for_timeout(3000)

        # Wait for data to load
        try:
            await page.wait_for_function(
                "document.querySelector('#total_traffic_card_idle [data-tag_type=\"total\"]')?.innerText?.length > 0",
                timeout=5000,
            )
        except Exception:
            pass

        # Extract portal metrics using correct card IDs
        data = await page.evaluate("""
            () => {
                function parseValue(text) {
                    if (!text) return 0;
                    let value = text.trim().replace(/,/g, '');
                    if (value.endsWith('K')) return parseFloat(value) * 1000;
                    if (value.endsWith('M')) return parseFloat(value) * 1000000;
                    return parseFloat(value) || 0;
                }
                
                const result = { pageview: 0, session: 0, new_users: 0, total_users: 0 };
                
                // PAGEVIEW = total_traffic_card_idle
                const pvCard = document.querySelector('#total_traffic_card_idle');
                if (pvCard) {
                    const el = pvCard.querySelector('[data-tag_type="total"]') || pvCard.querySelector('.fw-bold');
                    if (el) result.pageview = parseValue(el.innerText);
                }
                
                // SESSION = total_session_card_idle
                const sessCard = document.querySelector('#total_session_card_idle');
                if (sessCard) {
                    const el = sessCard.querySelector('[data-tag_type="total"]') || sessCard.querySelector('.fw-bold');
                    if (el) result.session = parseValue(el.innerText);
                }
                
                // NEW USERS = total_new_user_card_idle
                const newCard = document.querySelector('#total_new_user_card_idle');
                if (newCard) {
                    const el = newCard.querySelector('[data-tag_type="total"]') || newCard.querySelector('.fw-bold');
                    if (el) result.new_users = parseValue(el.innerText);
                }
                
                // TOTAL USERS = total_user_card_idle
                const totalCard = document.querySelector('#total_user_card_idle');
                if (totalCard) {
                    const el = totalCard.querySelector('[data-tag_type="total"]') || totalCard.querySelector('.fw-bold');
                    if (el) result.total_users = parseValue(el.innerText);
                }
                
                return result;
            }
        """)

        # Calculate Returning Users = Total Users - New Users
        returning_users = max(0, data["total_users"] - data["new_users"])

        results.append(
            {
                "SATKER": satker_name,
                "PAGEVIEW": int(data["pageview"]),
                "SESSION": int(data["session"]),
                "NEW USERS": int(data["new_users"]),
                "RETURNING USERS": int(returning_users),
            }
        )

        if (idx + 1) % 10 == 0:
            print(f"   Progress: {idx + 1}/{len(satker_options)} satkers done")

    return results


async def scrape_streaming_data(page, start_date, end_date):
    """Scrape Streaming data for all satkers."""
    print("\n" + "=" * 60)
    print("[*] STREAMING DATA SCRAPER")
    print("=" * 60)

    await page.goto(STREAMING_URL)
    await page.wait_for_load_state("networkidle")
    await page.wait_for_timeout(2000)

    satker_options = await get_satker_options(page)
    print(f"[*] Found {len(satker_options)} satkers to scrape")

    await set_date_range(page, start_date, end_date)

    results = []
    for idx, satker in enumerate(satker_options):
        satker_name = satker["text"]
        satker_value = satker["value"]

        print(f"[{idx + 1}/{len(satker_options)}] Scraping {satker_name}...")

        await page.select_option("#kt_dash_satker", value=satker_value)
        await page.click("#kt_apply_filter")
        await page.wait_for_timeout(3000)

        # Wait for data to load
        try:
            await page.wait_for_function(
                "document.querySelector('#total_hits_card_idle [data-tag_type=\"total\"]')?.innerText?.length > 0",
                timeout=5000,
            )
        except Exception:
            pass

        # Extract streaming metrics
        data = await page.evaluate("""
            () => {
                function parseValue(text) {
                    if (!text) return 0;
                    let value = text.trim().replace(/,/g, '');
                    if (value.endsWith('K')) return parseFloat(value) * 1000;
                    if (value.endsWith('M')) return parseFloat(value) * 1000000;
                    return parseFloat(value) || 0;
                }
                
                const result = {
                    pro1: 0, pro2: 0, pro3: 0, pro4: 0,
                    total_hits: 0, total_users: 0
                };
                
                // PRO 1-4
                ['pro1', 'pro2', 'pro3', 'pro4'].forEach(pro => {
                    const card = document.querySelector(`#${pro}_total_card_chart_line`);
                    if (card) {
                        const el = card.querySelector('[data-tag_type="total"]') || card.querySelector('.fs-4.text-gray-900.fw-bold');
                        if (el) result[pro] = parseValue(el.innerText);
                    }
                });
                
                // Total Hits
                const hitsCard = document.querySelector('#total_hits_card_idle');
                if (hitsCard) {
                    const el = hitsCard.querySelector('[data-tag_type="total"]') || hitsCard.querySelector('.fw-bold');
                    if (el) result.total_hits = parseValue(el.innerText);
                }
                
                // Total Users
                const usersCard = document.querySelector('#total_users_card_idle');
                if (usersCard) {
                    const el = usersCard.querySelector('[data-tag_type="total"]') || usersCard.querySelector('.fw-bold');
                    if (el) result.total_users = parseValue(el.innerText);
                }
                
                return result;
            }
        """)

        all_pro = data["pro1"] + data["pro2"] + data["pro3"] + data["pro4"]

        results.append(
            {
                "SATKER": satker_name,
                "PRO 1": int(data["pro1"]),
                "PRO 2": int(data["pro2"]),
                "PRO 3": int(data["pro3"]),
                "PRO 4": int(data["pro4"]),
                "ALL PRO": int(all_pro),
                "TOTAL HITS *DAP VER.": int(data["total_hits"]),
                "TOTAL USER": int(data["total_users"]),
            }
        )

        if (idx + 1) % 10 == 0:
            print(f"   Progress: {idx + 1}/{len(satker_options)} satkers done")

    return results


def export_portal_excel(data, start_date, end_date):
    """Export Portal data to Excel with template format and rankings."""
    df = pd.DataFrame(data)
    df = df.sort_values("SATKER")

    # Calculate rankings (higher value = better rank = lower number)
    df["RANK (PAGEVIEWS)"] = (
        df["PAGEVIEW"].rank(ascending=False, method="min").astype(int)
    )
    df["RANK (SESSION)"] = df["SESSION"].rank(ascending=False, method="min").astype(int)

    # Reorder columns to match template
    columns = [
        "SATKER",
        "PAGEVIEW",
        "SESSION",
        "NEW USERS",
        "RETURNING USERS",
        "RANK (PAGEVIEWS)",
        "RANK (SESSION)",
    ]
    df = df[columns]

    # Format date for filename (e.g., "JANUARI 2026")
    month_names = {
        1: "JANUARI",
        2: "FEBRUARI",
        3: "MARET",
        4: "APRIL",
        5: "MEI",
        6: "JUNI",
        7: "JULI",
        8: "AGUSTUS",
        9: "SEPTEMBER",
        10: "OKTOBER",
        11: "NOVEMBER",
        12: "DESEMBER",
    }
    start_dt = datetime.strptime(start_date, "%d/%m/%Y")
    month_name = month_names.get(start_dt.month, start_dt.strftime("%B").upper())
    output_file = f"{month_name} {start_dt.year}-Portal.xlsx"

    # Write to Excel with styled header
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Sheet1", index=False)

        # Style the header row (optional: add colors)
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        # Define fills for each column
        fills = [
            PatternFill(
                start_color="E06666", end_color="E06666", fill_type="solid"
            ),  # SATKER - Red
            PatternFill(
                start_color="93C47D", end_color="93C47D", fill_type="solid"
            ),  # PAGEVIEW - Green
            PatternFill(
                start_color="FFD966", end_color="FFD966", fill_type="solid"
            ),  # SESSION - Yellow
            PatternFill(
                start_color="6FA8DC", end_color="6FA8DC", fill_type="solid"
            ),  # NEW USERS - Blue
            PatternFill(
                start_color="8E7CC3", end_color="8E7CC3", fill_type="solid"
            ),  # RETURNING - Purple
            PatternFill(
                start_color="F6B26B", end_color="F6B26B", fill_type="solid"
            ),  # RANK PV - Orange
            PatternFill(
                start_color="76A5AF", end_color="76A5AF", fill_type="solid"
            ),  # RANK SESS - Teal
        ]

        for col_idx, fill in enumerate(fills, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    print(f"\n[+] Portal data exported to: {output_file}")
    print(f"[*] Summary:")
    print(f"   Total satkers: {len(df)}")
    print(f"   Satkers with pageviews: {(df['PAGEVIEW'] > 0).sum()}")
    print(f"   Total pageviews: {df['PAGEVIEW'].sum():,}")
    print(f"   Total sessions: {df['SESSION'].sum():,}")

    return output_file


def export_streaming_excel(data, start_date, end_date):
    """Export Streaming data to Excel with template format."""
    df = pd.DataFrame(data)
    df = df.sort_values("SATKER")

    # Reorder columns to match template
    columns = [
        "SATKER",
        "PRO 1",
        "PRO 2",
        "PRO 3",
        "PRO 4",
        "ALL PRO",
        "TOTAL HITS *DAP VER.",
        "TOTAL USER",
    ]
    df = df[columns]

    # Format date for filename
    month_names = {
        1: "JANUARI",
        2: "FEBRUARI",
        3: "MARET",
        4: "APRIL",
        5: "MEI",
        6: "JUNI",
        7: "JULI",
        8: "AGUSTUS",
        9: "SEPTEMBER",
        10: "OKTOBER",
        11: "NOVEMBER",
        12: "DESEMBER",
    }
    start_dt = datetime.strptime(start_date, "%d/%m/%Y")
    month_name = month_names.get(start_dt.month, start_dt.strftime("%B").upper())
    output_file = f"{month_name} {start_dt.year}-Streaming.xlsx"

    # Write with multi-level header matching template
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Sheet1", startrow=2, index=False, header=False)

        worksheet = writer.sheets["Sheet1"]

        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        # Row 1: Main headers
        main_headers = [
            ("SATKER", 1, 1, "CC9999"),  # Pink/salmon
            ("HITS", 2, 6, "F6B26B"),  # Orange (spans 5 columns)
            ("TOTAL HITS *DAP VER.", 7, 7, "6FA8DC"),  # Blue
            ("TOTAL USER", 8, 8, "93C47D"),  # Green
        ]

        for text, start_col, end_col, color in main_headers:
            cell = worksheet.cell(row=1, column=start_col, value=text)
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if start_col != end_col:
                worksheet.merge_cells(
                    start_row=1, start_column=start_col, end_row=1, end_column=end_col
                )

        # Row 2: Sub headers for HITS
        sub_headers = ["", "PRO 1", "PRO 2", "PRO 3", "PRO 4", "ALL PRO", "", ""]
        for col, text in enumerate(sub_headers, 1):
            if text:
                cell = worksheet.cell(row=2, column=col, value=text)
                cell.fill = PatternFill(
                    start_color="FFE599", end_color="FFE599", fill_type="solid"
                )  # Yellow
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        # Merge SATKER row 1-2
        worksheet.merge_cells("A1:A2")
        # Merge TOTAL HITS row 1-2
        worksheet.merge_cells("G1:G2")
        # Merge TOTAL USER row 1-2
        worksheet.merge_cells("H1:H2")

    print(f"\n[+] Streaming data exported to: {output_file}")
    print(f"[*] Summary:")
    print(f"   Total satkers: {len(df)}")
    print(f"   Satkers with data: {(df['ALL PRO'] > 0).sum()}")
    print(f"   Total ALL PRO hits: {df['ALL PRO'].sum():,}")
    print(f"   Total HITS (DAP VER.): {df['TOTAL HITS *DAP VER.'].sum():,}")
    print(f"   Total users: {df['TOTAL USER'].sum():,}")

    return output_file


async def main():
    parser = argparse.ArgumentParser(description="DAP RRI Unified Scraper")
    parser.add_argument(
        "--start-date", default="01/01/2026", help="Start date (DD/MM/YYYY)"
    )
    parser.add_argument(
        "--end-date", default="31/01/2026", help="End date (DD/MM/YYYY)"
    )
    parser.add_argument(
        "--portal-only", action="store_true", help="Scrape only Portal data"
    )
    parser.add_argument(
        "--streaming-only", action="store_true", help="Scrape only Streaming data"
    )
    args = parser.parse_args()

    scrape_portal = not args.streaming_only
    scrape_streaming = not args.portal_only

    print("=" * 60)
    print("[*] DAP RRI UNIFIED SCRAPER")
    print("=" * 60)
    print(f"[*] Date range: {args.start_date} - {args.end_date}")
    print(f"[*] Portal: {'Yes' if scrape_portal else 'No'}")
    print(f"[*] Streaming: {'Yes' if scrape_streaming else 'No'}")
    print()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()

        await login(page)

        portal_file = None
        streaming_file = None

        if scrape_portal:
            portal_data = await scrape_portal_data(page, args.start_date, args.end_date)
            if portal_data:
                portal_file = export_portal_excel(
                    portal_data, args.start_date, args.end_date
                )

        if scrape_streaming:
            streaming_data = await scrape_streaming_data(
                page, args.start_date, args.end_date
            )
            if streaming_data:
                streaming_file = export_streaming_excel(
                    streaming_data, args.start_date, args.end_date
                )

        await browser.close()

    print("\n" + "=" * 60)
    print("[+] SCRAPING COMPLETE!")
    print("=" * 60)
    if portal_file:
        print(f"[*] Portal: {portal_file}")
    if streaming_file:
        print(f"[*] Streaming: {streaming_file}")


if __name__ == "__main__":
    asyncio.run(main())
